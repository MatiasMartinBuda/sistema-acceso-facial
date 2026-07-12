"""
Módulo de Reportes Actualizado:
Muestra el dashboard en consola y genera un archivo Excel (.xlsx) corregido
dentro de la carpeta dedicada 'Reportes'.

Uso: python reporte.py
"""
import database
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os  # Para manejar la creación de carpetas

def exportar_a_excel_nativo():
    # 1. Aseguramos que exista la carpeta 'Reportes'
    carpeta_reportes = "Reportes"
    os.makedirs(carpeta_reportes, exist_ok=True)
    
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"Reporte_Accesos_{fecha_str}.xlsx"
    
    # Armamos la ruta completa para guardarlo adentro de la carpeta
    ruta_completa = os.path.join(carpeta_reportes, nombre_archivo)
    
    # 2. Obtener datos desde database.py
    r = database.reporte_semanal()
    personas = database.listar_personas()
    with database.get_conn() as conn:
        logs = conn.execute("SELECT * FROM logs_acceso ORDER BY timestamp DESC").fetchall()

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    
    # --- PESTAÑA 1: RESUMEN ---
    ws1 = wb.active
    ws1.title = "Resumen General"
    ws1.append(["Métrica", "Cantidad"])
    ws1.append(["Total de eventos", r['total']])
    ws1.append(["Accesos permitidos", r['permitidos']])
    ws1.append(["Accesos denegados", r['denegados']])
    ws1.append([])
    ws1.append(["Camino", "Cantidad"])
    for camino, cant in r["por_camino"].items():
        nombre_camino = f"Camino {camino} (Reconocido)" if camino == 'A' else f"Camino {camino} (PIN)" if camino == 'B' else f"Camino {camino} (Visita)"
        ws1.append([nombre_camino, cant])

    # --- PESTAÑA 2: HISTORIAL DETALLADO ---
    ws2 = wb.create_sheet(title="Historial de Accesos")
    ws2.append(["ID Log", "Fecha y Hora", "ID Persona", "DNI Declarado", "Depto Destino", "Camino", "Resultado", "Score", "Detalle"])
    for log in logs:
        ws2.append([
            log['id'], 
            str(log['timestamp']), 
            log['persona_id'] if log['persona_id'] is not None else "", 
            str(log['dni_declarado']) if log['dni_declarado'] is not None else "", 
            str(log['depto_destino']) if log['depto_destino'] is not None else "", 
            str(log['camino']), 
            str(log['resultado']), 
            log['score'] if log['score'] is not None else "", 
            str(log['detalle'])
        ])

    # --- PESTAÑA 3: PADRÓN DE PERSONAS ---
    ws3 = wb.create_sheet(title="Personas Registradas")
    ws3.append(["ID Modelo LBPH", "Apellido", "Nombre", "Categoría", "Departamento", "Tipo Acceso", "Lista Negra"])
    for p in personas:
        es_lista_negra = "SÍ" if p["lista_negra"] else "NO"
        ws3.append([
            p['label_lbph'], 
            str(p['apellido']), 
            str(p['nombre']), 
            str(p['categoria']), 
            str(p['depto']) if p['depto'] is not None else "", 
            str(p['tipo_acceso']), 
            es_lista_negra
        ])

    # 4. AUTOAJUSTAR COLUMNAS
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 5. Guardar archivo en la nueva ruta
    try:
        wb.save(ruta_completa)
        print(f"\n[OK] ¡Reporte exportado con éxito! Guardado en: {ruta_completa}")
    except Exception as e:
        print(f"\n[ERROR] No se pudo guardar el archivo Excel: {e}")

def main():
    database.init_db()
    r = database.reporte_semanal()

    print("=== TABLERO DE CONTROL: Últimos 7 días ===")
    print(f"Total de eventos:     {r['total']}")
    print(f"Accesos permitidos:   {r['permitidos']}")
    print(f"Accesos denegados:    {r['denegados']}")
    print("Por camino (A=reconocido, B=PIN residente, C=visita):")
    for camino, cant in r["por_camino"].items():
        print(f"   Camino {camino}: {cant}")

    print("\n=== Personas registradas ===")
    for p in database.listar_personas():
        negra = " [LISTA NEGRA]" if p["lista_negra"] else ""
        print(f"  #{p['label_lbph']:>3} {p['apellido']}, {p['nombre']} "
              f"- {p['categoria']} - depto {p['depto']}{negra}")
              
    exportar_a_excel_nativo()

if __name__ == "__main__":
    main()